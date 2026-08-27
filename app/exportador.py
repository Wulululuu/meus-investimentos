"""Exporta a carteira (posicoes, compras e proventos) para um arquivo Excel,
util para declaracao de imposto de renda ou planilhas proprias."""
from __future__ import annotations

import datetime as dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from .database import get_conn

EXPORTS_DIR = Path(__file__).resolve().parent.parent / "exports"


def _cabecalho(ws, colunas: list[str]) -> None:
    ws.append(colunas)
    for cel in ws[1]:
        cel.font = Font(bold=True)


def exportar_carteira() -> Path:
    conn = get_conn()
    try:
        lotes = conn.execute(
            "SELECT * FROM investimentos ORDER BY ticker, data_compra"
        ).fetchall()
        cotacoes = {
            r["ticker"]: r for r in conn.execute("SELECT * FROM cotacoes_atuais").fetchall()
        }
        proventos = conn.execute(
            "SELECT * FROM proventos_recebidos ORDER BY ticker, data_ex"
        ).fetchall()
        vendas = conn.execute(
            "SELECT * FROM vendas ORDER BY ticker, data_venda"
        ).fetchall()
    finally:
        conn.close()

    por_ticker: dict[str, list] = {}
    for lote in lotes:
        por_ticker.setdefault(lote["ticker"], []).append(lote)

    wb = Workbook()

    ws_posicoes = wb.active
    ws_posicoes.title = "Posições"
    _cabecalho(ws_posicoes, [
        "Ticker", "Tipo", "Quantidade", "Preço médio", "Valor investido",
        "Preço atual", "Valorização", "Proventos recebidos", "Saldo total",
    ])
    for ticker, lotes_ticker in sorted(por_ticker.items()):
        tipo = lotes_ticker[0]["tipo"]
        quantidade = sum(l["quantidade"] for l in lotes_ticker)
        custo_total = sum(l["quantidade"] * l["preco_medio_compra"] for l in lotes_ticker)
        preco_medio = custo_total / quantidade if quantidade else 0
        cot = cotacoes.get(ticker)
        preco_atual = cot["preco_atual"] if cot else None

        prov_ticker = [p for p in proventos if p["ticker"] == ticker]
        proventos_total = 0.0
        for lote in lotes_ticker:
            proventos_total += sum(
                p["valor_por_cota"] for p in prov_ticker if p["data_ex"] >= lote["data_compra"]
            ) * lote["quantidade"]

        valorizacao = (preco_atual * quantidade - custo_total) if preco_atual is not None else None
        saldo_total = (valorizacao + proventos_total) if valorizacao is not None else None

        ws_posicoes.append([
            ticker, tipo, quantidade, round(preco_medio, 4), round(custo_total, 2),
            preco_atual, valorizacao, round(proventos_total, 2), saldo_total,
        ])

    ws_compras = wb.create_sheet("Compras")
    _cabecalho(ws_compras, ["Ticker", "Tipo", "Data da compra", "Quantidade", "Preço pago", "Valor total"])
    for lote in lotes:
        ws_compras.append([
            lote["ticker"], lote["tipo"], lote["data_compra"], lote["quantidade"],
            lote["preco_medio_compra"], round(lote["quantidade"] * lote["preco_medio_compra"], 2),
        ])

    ws_vendas = wb.create_sheet("Vendas")
    _cabecalho(ws_vendas, ["Ticker", "Data da venda", "Quantidade", "Preço unitário", "Valor total"])
    for v in vendas:
        ws_vendas.append([
            v["ticker"], v["data_venda"], v["quantidade"], v["preco_unitario"],
            round(v["quantidade"] * v["preco_unitario"], 2),
        ])

    ws_proventos = wb.create_sheet("Proventos recebidos")
    _cabecalho(ws_proventos, ["Ticker", "Data (ex)", "Valor por cota"])
    for p in proventos:
        ws_proventos.append([p["ticker"], p["data_ex"], p["valor_por_cota"]])

    EXPORTS_DIR.mkdir(exist_ok=True)
    nome_arquivo = f"carteira_{dt.date.today().isoformat()}.xlsx"
    caminho = EXPORTS_DIR / nome_arquivo
    wb.save(caminho)
    return caminho
